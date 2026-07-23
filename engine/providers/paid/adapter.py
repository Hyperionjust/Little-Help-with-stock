"""声明式适配器引擎（方案 §4.2，阶段2-1）。

为什么用 YAML 而非每源一个 Python 文件：
  "任意付费终端有 key 即可接入"这个目标下，大多数终端的差异只在
  端点、认证、字段名、单位——不在逻辑。把这些差异声明进 YAML，
  一个引擎解释所有适配器，加一家终端只加一个 YAML，不碰 Python。

一个适配器 = 一份 YAML（见 adapters/*.yaml，custom.template.yaml 是带注释的模板）。
引擎读它、按 transport 取数、按 fields 映射并归一化单位，产出 base.Provider 接口。

四种 transport：
  http           REST/XHR 端点（tushare_pro、bloomberg BQL 服务、自建数据服务）
  file_drop      用户上传的导出文件（Wind/iFind/同花顺/Choice 导出的 xlsx·csv）
  cli            本地安装的终端客户端（Wind Python 接口、彭博本地终端）
  host_function  宿主注入的连接器（运行环境提供数据连接器时）

单位纪律：适配器只声明字段的**原始单位**，换算一律走 _util.normalize_unit——
与 §4.3 集中归一化对齐，量纲哨兵作为最后防线。
"""
from __future__ import annotations
import os
import sys

_HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(_HERE, "..", ".."))
import _paths  # noqa: F401,E402
from _util import normalize_unit  # noqa: E402
from base import Provider, QUOTE_FIELDS, FINANCIAL_FIELDS  # noqa: E402

_MARKET_INTERNAL = "million"  # engine 内部统一口径（INTERNAL_UNIT 的英文别名）


def _num(x):
    try:
        f = float(x)
        return None if f != f else f
    except (TypeError, ValueError):
        return None


class AdapterError(RuntimeError):
    pass


class DeclarativeProvider(Provider):
    """由一份 YAML 规格驱动的 provider。行为全部来自 spec，无硬编码逻辑。"""

    def __init__(self, spec, transport=None):
        self.spec = spec
        self.name = spec["name"]
        self.display_name = spec.get("display_name", self.name)
        self.supports_markets = tuple(spec.get("markets", []))
        self.tier = spec.get("tier", 1)
        self._transport = transport or _make_transport(spec)

    # ── 能力探测（registry 用）─────────────────────────────────────
    def available(self):
        """认证是否就绪。不做网络探测——那是 registry.health_probe 的职责。"""
        auth = self.spec.get("auth", {})
        method = auth.get("method", "none")
        if method == "none":
            return True
        if method == "env_token":
            return bool(os.environ.get(auth.get("env_var", "")))
        if method == "file_drop":
            return bool(os.environ.get(auth.get("path_env", "ERS_IMPORT_FILE")))
        if method in ("cli", "host_function", "local_terminal"):
            # 由 transport 自证；registry 会做健康探测
            return self._transport.probe_available()
        return False

    def capabilities(self):
        return set(self.spec.get("capabilities", []))

    # ── 数据接口（映射 + 归一化）───────────────────────────────────
    # 每股类字段是原币计价、不做规模换算（价格就是价格）。其余货币金额统一到"百万"。
    _NO_SCALE_FIELDS = {"price", "prev_close", "eps_basic", "eps_diluted"}

    def _map_group(self, group, raw, field_whitelist):
        """按 spec.fields[group] 把上游原始记录映射成标准化字段，并归一化单位。"""
        fmap = (self.spec.get("fields", {}) or {}).get(group, {})
        out = {}
        for std_field, rule in fmap.items():
            # 只填标准字段清单内的键（防 YAML 写错字段污染下游）
            top = std_field.split(".")[0]
            if field_whitelist and top not in field_whitelist and std_field not in field_whitelist:
                continue
            val = _extract(raw, rule.get("path", std_field))
            num = _num(val)
            unit = rule.get("unit")
            need_scale = unit and unit not in ("native", "raw") and std_field not in self._NO_SCALE_FIELDS
            if num is not None and need_scale:
                try:
                    num = normalize_unit(num, unit, _MARKET_INTERNAL)
                except ValueError as e:
                    raise AdapterError(f"{self.name}.{group}.{std_field}: {e}") from None
            out[std_field] = num if num is not None else val
        return out

    def get_quote(self, resolved):
        raw = self._transport.fetch("quote", resolved)
        if not raw:
            return {}
        out = {k: None for k in QUOTE_FIELDS}
        out.update(self._map_group("quote", raw, QUOTE_FIELDS))
        return out

    def get_financials(self, resolved):
        raw = self._transport.fetch("financials", resolved)
        if not raw:
            return {}
        periods = raw if isinstance(raw, list) else raw.get("periods", [])
        annual = []
        for rec in periods:
            row = self._map_group("financials", rec, FINANCIAL_FIELDS)
            row["period"] = rec.get("period") or rec.get("report_date") or rec.get("date")
            annual.append(row)
        return {"source": self.name,
                "accounting_standard": self.spec.get("accounting_standard"),
                "annual": [r for r in annual if r.get("period")]}

    def get_kline(self, resolved, adjust="qfq"):
        raw = self._transport.fetch("kline", resolved, adjust=adjust)
        return raw or {}

    def get_estimates(self, resolved):
        raw = self._transport.fetch("estimates", resolved)
        if not raw:
            return {}
        return self._map_group("estimates", raw, None)


# ── 字段路径提取（支持点路径 a.b.c 与列表首元素）─────────────────────
def _extract(obj, path):
    cur = obj
    for part in str(path).split("."):
        if isinstance(cur, list):
            cur = cur[0] if cur else None
        if isinstance(cur, dict):
            cur = cur.get(part)
        else:
            return None
    return cur


# ── transport 工厂 ──────────────────────────────────────────────────
def _make_transport(spec):
    kind = (spec.get("transport", {}) or {}).get("kind", "http")
    if kind == "http":
        return HttpTransport(spec)
    if kind == "file_drop":
        return FileDropTransport(spec)
    if kind in ("cli", "host_function", "local_terminal"):
        return StubTransport(spec, kind)
    raise AdapterError(f"未知 transport.kind: {kind}")


class _Transport:
    def __init__(self, spec):
        self.spec = spec

    def probe_available(self):
        return False

    def fetch(self, capability, resolved, **kw):
        return None


class HttpTransport(_Transport):
    """REST/XHR。真实厂商实现留给用户按 vendor 文档补全 endpoint 细节——
    引擎负责认证注入、超时重试、响应定位，这些是通用的。"""

    def probe_available(self):
        return True

    def fetch(self, capability, resolved, **kw):
        import urllib.request
        import json as _json
        t = self.spec.get("transport", {})
        eps = t.get("endpoints", {})
        ep = eps.get(capability)
        if not ep:
            return None
        url = _render(ep.get("url", ""), resolved, self.spec)
        headers = {k: _render(v, resolved, self.spec) for k, v in (ep.get("headers") or {}).items()}
        timeout = t.get("timeout_s", 20)
        req = urllib.request.Request(url, headers=headers)
        try:
            with urllib.request.urlopen(req, timeout=timeout) as r:
                data = _json.loads(r.read().decode("utf-8"))
        except Exception:
            return None
        root = ep.get("response_root")
        return _extract(data, root) if root else data


class FileDropTransport(_Transport):
    """用户上传的导出文件（统一 import_local 的能力，方案 §12.1）。

    路径由环境变量 ERS_IMPORT_FILE 指定（run_analysis 在用户提供文件时设置）。
    别名映射与单位声明来自 YAML 的 fields，与其它 transport 一致。
    """

    def probe_available(self):
        return bool(os.environ.get(self.spec.get("auth", {}).get("path_env", "ERS_IMPORT_FILE")))

    def fetch(self, capability, resolved, **kw):
        path = os.environ.get(self.spec.get("auth", {}).get("path_env", "ERS_IMPORT_FILE"))
        if not path or not os.path.exists(path):
            return None
        rows = _read_tabular(path)
        if capability == "financials":
            return _rows_to_periods(rows, self.spec)
        if capability == "quote":
            return rows[0] if rows else None
        return None


class StubTransport(_Transport):
    """cli / host_function / local_terminal 的占位。

    这些依赖运行环境或本地安装，开发期无法真实验证。probe 默认 False，
    有相应环境时由用户在 YAML 的 auth 里声明探测方式。真实对接时替换本类。
    """

    def __init__(self, spec, kind):
        super().__init__(spec)
        self.kind = kind

    def probe_available(self):
        env = self.spec.get("auth", {}).get("probe_env")
        return bool(os.environ.get(env)) if env else False


# ── 模板渲染：{symbol} {token} {env:VAR} 占位 ────────────────────────
def _render(template, resolved, spec):
    if not isinstance(template, str):
        return template
    s = template
    for k, v in (resolved or {}).items():
        s = s.replace("{%s}" % k, str(v))
    import re
    for m in re.findall(r"\{env:([A-Z_][A-Z0-9_]*)\}", s):
        s = s.replace("{env:%s}" % m, os.environ.get(m, ""))
    auth = spec.get("auth", {})
    if auth.get("env_var"):
        s = s.replace("{token}", os.environ.get(auth["env_var"], ""))
    return s


# ── 导出文件读取 + 别名映射 ─────────────────────────────────────────
def _read_tabular(path):
    """xlsx/csv → list[dict]。列名保持原样，映射交给 fields 的 aliases。"""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls"):
        try:
            from openpyxl import load_workbook
        except ImportError:
            return []
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        header = [str(c) if c is not None else "" for c in rows[0]]
        return [dict(zip(header, r)) for r in rows[1:]]
    if ext in (".csv", ".tsv"):
        import csv
        delim = "\t" if ext == ".tsv" else ","
        with open(path, encoding="utf-8-sig") as f:
            return list(csv.DictReader(f, delimiter=delim))
    return []


def _rows_to_periods(rows, spec):
    """把导出表的行整理成 [{period, <alias-resolved fields>}]。

    导出文件常是"字段为行、期间为列"的转置布局，也可能是"期间为行"。
    这里用 fields.financials 的 aliases 做列名解析，两种布局都尝试。
    """
    fmap = (spec.get("fields", {}) or {}).get("financials", {})
    alias_to_std = {}
    for std, rule in fmap.items():
        for a in rule.get("aliases", []):
            alias_to_std[str(a).strip()] = std

    # 布局A：每行一个期间（有 period/报告期 列）
    period_keys = {"period", "报告期", "report_date", "date", "年度", "会计期间"}
    if rows and any(k in rows[0] for k in period_keys):
        out = []
        for r in rows:
            per = next((r[k] for k in period_keys if r.get(k)), None)
            if not per:
                continue
            rec = {"period": str(per)}
            for col, val in r.items():
                std = alias_to_std.get(str(col).strip())
                if std:
                    rec[std] = val
            out.append(rec)
        return out

    # 布局B：字段为行、期间为列（转置）
    if rows:
        name_col = next((c for c in rows[0] if str(c).strip() in
                         ({"科目", "项目", "指标", "item", "field"} | set(alias_to_std))), None)
        if name_col:
            periods = {}
            for r in rows:
                std = alias_to_std.get(str(r.get(name_col, "")).strip())
                if not std:
                    continue
                for col, val in r.items():
                    if col == name_col:
                        continue
                    periods.setdefault(str(col), {"period": str(col)})[std] = val
            return list(periods.values())
    return []
