"""渲染 Excel 底稿：Raw Data sheet 存原始数据，Ratios sheet 用**活公式**引用 Raw Data。

医药标的额外 rNPV sheet：PoS/峰值销售/折现率为可编辑输入格，rNPV 用 Excel 公式实时重算。
verify_consistency.py 会读公式串与计算值双份校验，确认是活公式而非硬编码。
用法：python render_xlsx.py analysis.json [-o workbook.xlsx]
"""
from __future__ import annotations
import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _util import load_json  # noqa: E402


def _v(node):
    return node.get("value") if isinstance(node, dict) else node


def render(analysis, out):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    # ---------- Raw Data ----------
    raw = wb.active
    raw.title = "Raw Data"
    raw["A1"] = "字段"; raw["B1"] = "值"; raw["C1"] = "数据源"; raw["D1"] = "口径/公式"
    for c in "ABCD":
        raw[f"{c}1"].font = Font(bold=True)

    fin = analysis.get("_raw_financials") or {}
    # 从 analysis 取关键原始量（annotated value 的 value + 我们额外从 quote/fin 落原始数）
    q = analysis.get("quote", {})
    # 关键原始输入放入 Raw Data，供 Ratios 公式引用
    rows = [
        ("price", _v(q.get("price")), (q.get("price") or {}).get("source", "")),
        ("market_cap", _v(q.get("market_cap")), (q.get("market_cap") or {}).get("source", "")),
    ]
    # 财务原始（从 valuation/profitability 的分母还原不可靠，改从 meta._raw 落；若无则用比率反填说明）
    raw_fin = analysis.get("meta", {}).get("_raw_inputs", {})
    for k in ("revenue_ttm", "net_income_ttm", "equity", "total_assets", "total_liabilities",
              "current_assets", "current_liabilities", "inventory", "gross_profit", "ocf"):
        rows.append((k, raw_fin.get(k), "financials"))
    r = 2
    ref = {}
    for name, val, src in rows:
        raw[f"A{r}"] = name
        raw[f"B{r}"] = val
        raw[f"C{r}"] = src
        ref[name] = f"'Raw Data'!B{r}"
        r += 1

    # ---------- Ratios（活公式）----------
    rat = wb.create_sheet("Ratios")
    rat["A1"] = "比率"; rat["B1"] = "公式值(活)"; rat["C1"] = "口径"
    for c in "ABC":
        rat[f"{c}1"].font = Font(bold=True)
    formulas = [
        ("PE-TTM", f"={ref['market_cap']}/{ref['net_income_ttm']}", "市值/归母净利TTM"),
        ("PB", f"={ref['market_cap']}/{ref['equity']}", "市值/归母净资产"),
        ("PS-TTM", f"={ref['market_cap']}/{ref['revenue_ttm']}", "市值/营收TTM"),
        ("净利率", f"={ref['net_income_ttm']}/{ref['revenue_ttm']}", "归母净利/营收"),
        ("毛利率", f"={ref['gross_profit']}/{ref['revenue_ttm']}", "毛利/营收"),
        ("资产负债率", f"={ref['total_liabilities']}/{ref['total_assets']}", "总负债/总资产"),
        ("流动比率", f"={ref['current_assets']}/{ref['current_liabilities']}", "流动资产/流动负债"),
        ("速动比率", f"=({ref['current_assets']}-{ref['inventory']})/{ref['current_liabilities']}",
         "(流动资产-存货)/流动负债"),
        ("现金含量", f"={ref['ocf']}/{ref['net_income_ttm']}", "OCF/净利"),
    ]
    rr = 2
    for name, fml, note in formulas:
        rat[f"A{rr}"] = name
        rat[f"B{rr}"] = fml   # 活公式串
        rat[f"C{rr}"] = note
        rr += 1

    # ---------- 医药 rNPV sheet ----------
    ph = analysis.get("pharma")
    if ph:
        rn = wb.create_sheet("rNPV")
        hdr = ["资产", "适应症", "当前阶段", "累积PoS(可改)", "峰值销售(可改)",
               "折现率(可改)", "净现金流现值系数", "rNPV(公式)"]
        for i, h in enumerate(hdr):
            cell = rn.cell(row=1, column=i + 1, value=h)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="FFE4B5")
        rr = 2
        for a in ph.get("rnpv", {}).get("assets", []):
            rn.cell(row=rr, column=1, value=a.get("asset"))
            rn.cell(row=rr, column=2, value=a.get("indication"))
            rn.cell(row=rr, column=3, value=a.get("current_phase"))
            rn.cell(row=rr, column=4, value=_v(a.get("cumulative_pos")))   # 可编辑输入
            rn.cell(row=rr, column=5, value=_v(a.get("peak_sales")))       # 可编辑输入
            rn.cell(row=rr, column=6, value=_v(ph.get("clinical_discount_rate")))  # 可编辑
            rn.cell(row=rr, column=7, value=3.5)  # 简化的现金流现值系数占位（可改）
            # rNPV 活公式：峰值销售 × 利润率(0.35) × 累积PoS × 现值系数
            rn.cell(row=rr, column=8,
                    value=f"=E{rr}*0.35*D{rr}*G{rr}")
            rr += 1
        # 公司层
        rn.cell(row=rr + 1, column=1, value="公司层 rNPV = ΣrNPV + 净现金 - 债务").font = Font(bold=True)
        rn.cell(row=rr + 1, column=8, value=f"=SUM(H2:H{rr-1})+{ph.get('rnpv',{}).get('net_cash',{}).get('value',0) or 0}-{ph.get('rnpv',{}).get('debt',{}).get('value',0) or 0}")
        # 核对清单 sheet
        cl = wb.create_sheet("需人工核对清单")
        cl["A1"] = "项目"; cl["B1"] = "取值"; cl["C1"] = "来源类型"; cl["D1"] = "对rNPV影响"
        for c in "ABCD":
            cl[f"{c}1"].font = Font(bold=True)
            cl[f"{c}1"].fill = PatternFill("solid", fgColor="F08080")
        cr = 2
        for item in ph.get("human_verification_checklist", []):
            cl.cell(row=cr, column=1, value=item.get("item"))
            v = item.get("value")
            cl.cell(row=cr, column=2, value=(_v(v) if isinstance(v, dict) else v))
            cl.cell(row=cr, column=3, value=item.get("source_type"))
            cl.cell(row=cr, column=4, value=item.get("impact_on_rnpv"))
            cr += 1

    # ---------- 质量门禁 sheet ----------
    g = wb.create_sheet("质量门禁")
    g["A1"] = "级别"; g["B1"] = "检查项"; g["C1"] = "信息"
    for c in "ABC":
        g[f"{c}1"].font = Font(bold=True)
    gr = 2
    qr = analysis.get("quality_report", {})
    for c in qr.get("critical", []):
        g.cell(row=gr, column=1, value="CRITICAL"); g.cell(row=gr, column=2, value=c["check"])
        g.cell(row=gr, column=3, value=c["message"]); gr += 1
    for w in qr.get("warning", []):
        g.cell(row=gr, column=1, value="WARNING"); g.cell(row=gr, column=2, value=w["check"])
        g.cell(row=gr, column=3, value=w["message"]); gr += 1

    wb.save(out)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("analysis")
    ap.add_argument("-o", "--out")
    args = ap.parse_args()
    analysis = load_json(args.analysis)
    out = args.out or f"{analysis.get('resolution', {}).get('symbol', 'out')}_workbook.xlsx"
    render(analysis, out)
    print(out)


if __name__ == "__main__":
    main()
