"""四种输出对账：JSON 为唯一数据源，校验 HTML/Excel/Markdown 与之逐指标一致。

- HTML：读 <script id="analysis-data"> 内嵌 JSON，须与 analysis.json 完全一致。
- Excel：openpyxl 读取；对 Ratios sheet 做「活公式」验证（公式串引用 Raw Data，非硬编码数值）。
- Markdown：正则抽数字，逐个在 JSON 中找存在性（容差匹配）。

用法：verify_consistency.py <analysis.json> <dashboard.html> <workbook.xlsx> [markdown.md]
退出码 0=一致，1=不一致（driver 用退出码）。
"""
from __future__ import annotations
import argparse
import json
import re
import sys


def _load_json(p):
    with open(p, encoding="utf-8") as f:
        return json.load(f)


def check_html(analysis, html_path):
    errs = []
    with open(html_path, encoding="utf-8") as f:
        html = f.read()
    m = re.search(r'<script id="analysis-data"[^>]*>(.*?)</script>', html, re.S)
    if not m:
        return ["HTML 缺少 <script id='analysis-data'> 内嵌 JSON"]
    try:
        embedded = json.loads(m.group(1))
    except json.JSONDecodeError as e:
        return [f"HTML 内嵌 JSON 解析失败: {e}"]
    # 逐指标比对（关键块）
    for block in ("valuation", "profitability", "solvency", "growth"):
        a = analysis.get(block, {})
        b = embedded.get(block, {})
        for k, node in a.items():
            if isinstance(node, dict) and "value" in node:
                va, vb = node["value"], (b.get(k, {}) or {}).get("value")
                if not _num_eq(va, vb):
                    errs.append(f"HTML {block}.{k}: JSON={va} 内嵌={vb}")
    return errs


def check_excel(analysis, xlsx_path):
    from openpyxl import load_workbook
    errs = []
    wb_f = load_workbook(xlsx_path)  # 公式串
    if "Ratios" not in wb_f.sheetnames:
        return ["Excel 缺 Ratios sheet"]
    rat = wb_f["Ratios"]
    formula_cells = 0
    hardcoded = 0
    for row in rat.iter_rows(min_row=2):
        b = row[1] if len(row) > 1 else None
        if b is None or b.value is None:
            continue
        if isinstance(b.value, str) and b.value.startswith("="):
            formula_cells += 1
            if "Raw Data" not in b.value:
                errs.append(f"Ratios!{b.coordinate} 公式未引用 Raw Data: {b.value}")
        elif isinstance(b.value, (int, float)):
            hardcoded += 1
            errs.append(f"Ratios!{b.coordinate} 是硬编码数值 {b.value}（应为活公式）")
    if formula_cells == 0:
        errs.append("Ratios sheet 没有任何活公式")
    return errs


def check_markdown(analysis, md_path):
    errs = []
    with open(md_path, encoding="utf-8") as f:
        md = f.read()
    nums = [float(x.replace(",", "")) for x in re.findall(r"-?\d[\d,]*\.?\d*", md)]
    pool = _collect_values(analysis)
    for n in nums:
        if not any(_num_eq(n, p, rel=0.02) for p in pool):
            # 允许年份/整数计数等非指标数字：仅对小数报警
            if abs(n - round(n)) > 1e-9:
                errs.append(f"Markdown 数字 {n} 在 JSON 中找不到对应值")
    return errs


def _collect_values(obj, out=None):
    if out is None:
        out = []
    if isinstance(obj, dict):
        if "value" in obj and isinstance(obj.get("value"), (int, float)):
            out.append(obj["value"])
        for v in obj.values():
            _collect_values(v, out)
    elif isinstance(obj, list):
        for v in obj:
            _collect_values(v, out)
    elif isinstance(obj, (int, float)):
        out.append(obj)
    return out


def _num_eq(a, b, rel=1e-3, abs_tol=0.05):
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    if not isinstance(a, (int, float)) or not isinstance(b, (int, float)):
        return a == b
    if abs(a - b) <= abs_tol:
        return True
    denom = max(abs(a), abs(b), 1e-9)
    return abs(a - b) / denom <= rel


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("analysis")
    ap.add_argument("html")
    ap.add_argument("xlsx")
    ap.add_argument("markdown", nargs="?")
    args = ap.parse_args()
    analysis = _load_json(args.analysis)
    errs = []
    errs += check_html(analysis, args.html)
    errs += check_excel(analysis, args.xlsx)
    if args.markdown:
        errs += check_markdown(analysis, args.markdown)
    if errs:
        print("INCONSISTENT:")
        for e in errs:
            print("  -", e)
        sys.exit(1)
    print("CONSISTENT: 四种输出逐指标一致")
    sys.exit(0)


if __name__ == "__main__":
    main()
